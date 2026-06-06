import os, io
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image as PILImage

DST = '美甲分类_打标.xlsx'

# 款式图打标：(甲型, 手型, 长度, [款式], 颜色描述)
labels = [
    ('椭圆', '修长手', '短', ['纯色'], '裸杏色'),
    ('方形', '匀称手', '中', ['跳色', '纯色'], '橄榄绿+米白+焦糖'),
    ('方圆', '匀称手', '短', ['格纹', '跳色', '立体钻饰'], '黑色+粉红+棕色'),
    ('椭圆', '修长手', '中', ['镜面', '亮片', '纯色'], '香槟金+灰白'),
    ('方形', '匀称手', '短', ['豹纹', '纯色', '细边法式'], '奶白+黑色斑点+裸粉'),
    ('梯形', '修长手', '长', ['立体钻饰', '渐变', '亮片'], '裸粉+透明+香槟金'),
    ('尖头型', '修长手', '长', ['豹纹', '亮片'], '银色亮片+棕色豹纹'),
    ('尖头型', '修长手', '长', ['单条法式猫眼', '立体钻饰', '猫眼'], '黑色+裸透+银闪'),
    ('椭圆', '修长手', '长', ['手绘', '纯色'], '黑色+裸透'),
    ('梯形', '修长手', '长', ['手绘', '细边法式'], '裸粉+白色'),
    ('杏仁', '修长手', '长', ['贴纸', '手绘'], '裸透+白色樱花'),
    ('杏仁', '修长手', '长', ['手绘', '日系', '高位法式'], '裸透+黑色'),
    ('梯形', '修长手', '长', ['手绘', '立体钻饰'], '裸透+珠光+银色'),
    ('杏仁', '修长手', '长', ['单色渐变', '镜面'], '奶白+裸粉'),
    ('方形', '肉肉手', '短', ['跳色', '格纹', '手绘'], '橙红+酒红+雾粉+浅蓝+米色'),
    ('椭圆', '匀称手', '短', ['砂糖粉', '渐变'], '玫红+粉色+裸色'),
    ('梯形', '肉肉手', '长', ['立体钻饰', '珍珠', '细边法式'], '裸色+米白+珍珠白'),
    ('梯形', '修长手', '长', ['立体钻饰', '纯色'], '裸透+彩色钻'),
    ('杏仁', '修长手', '长', ['立体钻饰', '亮片', '豹纹'], '香槟金+棕色豹纹'),
    ('杏仁', '修长手', '长', ['豹纹', '立体钻饰', '亮片'], '裸粉+银色亮片+棕色豹纹'),
    ('方圆', '修长手', '短', ['金箔', '立体钻饰', '细边法式'], '裸色+金箔+蓝色+祖母绿'),
    ('梯形', '修长手', '中', ['镜面'], '玫瑰金'),
    ('杏仁', '修长手', '长', ['渐变', '立体钻饰'], '裸色+奶白'),
    ('圆形', '修长手', '短', ['亮片', '砂糖粉'], '透明+多彩亮片'),
    ('方形', '修长手', '中', ['渐变法式', '手绘'], '酒红+裸色'),
]

hand_labels = ['修长手','肉肉手','修长手','匀称手','骨节手','匀称手','修长手','修长手','肉肉手','骨节手','骨节手','肉肉手','修长手']

STYLE_OPTIONS = [
    '腮红渐变','对角渐变','三色渐变','空间渐变','反渐变','上下渐变','单色渐变','中间渐变','渐变',
    '宽型猫眼','微笑猫眼','法式猫眼','追光猫眼','经典猫眼','混夭绫猫眼','渐变猫眼','反渐变猫眼','弧形猫眼','十字猫眼','黑洞猫眼','单条法式猫眼','猫眼',
    '高位法式','标准法式','圆法式','平法式','爱心法式','交叉法式','斜法式','V型法式','细边法式','梯形法式','轮廓法式','圣诞法式','镂空法式','反渐变法式','渐变法式',
    '金箔','珍珠','链条','贝壳','砂糖粉','干花','极光纸','钢珠','转印纸','贴纸','平底钻','亮片','立体钻饰','金属片','尖底钴','晕染','手绘','镜面','纯色','磨砂','浮雕立体','日系','韩系','欧美','中式','主题','跳色','格纹','豹纹','魔镜粉'
]
NAIL_SHAPE = ['圆形','椭圆','方形','方圆','梯形','杏仁','尖头型']
HAND_TYPE = ['尖锥手','骨节手','修长手','匀称手','肉肉手','短粗手']
LENGTH = ['短','中','长']


def make_thumb_bytes(src_path, max_h=200):
    pil = PILImage.open(src_path).convert('RGB')
    w, h = pil.size
    if h > max_h:
        nw = int(w * max_h / h)
        pil = pil.resize((nw, max_h), PILImage.LANCZOS)
        w, h = nw, max_h
    bio = io.BytesIO()
    pil.save(bio, 'JPEG', quality=82, optimize=True)
    bio.seek(0)
    return bio, w, h


wb = load_workbook(DST)

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='5C7A99')
thin = Side(border_style='thin', color='BBBBBB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)


def add_image(ws, img_path, row):
    if not os.path.exists(img_path):
        return
    bio, w, h = make_thumb_bytes(img_path)
    target_h = 120
    scale = target_h / h
    xi = XLImage(bio)
    xi.width = int(w * scale)
    xi.height = target_h
    xi.anchor = f'B{row}'
    ws.add_image(xi)


def build_style_sheet(sheet_name, folder, labels):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    max_styles = max(len(l[3]) for l in labels)
    headers = ['序号', '图片', '甲型', '手型', '长度'] + ['款式'] * max_styles + ['颜色']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center
        c.border = border
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    for i in range(len(headers) - 2):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14
    ws.column_dimensions[get_column_letter(len(headers))].width = 30
    ws.row_dimensions[1].height = 28

    for n, (shape, hand, length, styles, color) in enumerate(labels, 1):
        r = n + 1
        ws.row_dimensions[r].height = 95
        ws.cell(row=r, column=1, value=f'{n:02d}').alignment = align_center
        ws.cell(row=r, column=1).border = border
        add_image(ws, os.path.join(folder, f'{n:02d}.png'), r)
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=3, value=shape).alignment = align_center
        ws.cell(row=r, column=4, value=hand).alignment = align_center
        ws.cell(row=r, column=5, value=length).alignment = align_center
        for ci in range(3, 6):
            ws.cell(row=r, column=ci).border = border
        for si in range(max_styles):
            v = styles[si] if si < len(styles) else ''
            c = ws.cell(row=r, column=6 + si, value=v)
            c.alignment = align_center
            c.border = border
        c = ws.cell(row=r, column=6 + max_styles, value=color)
        c.alignment = align_center
        c.border = border

    rows = len(labels)
    dv = DataValidation(type='list', formula1=f'"{",".join(NAIL_SHAPE)}"', allow_blank=True)
    dv.add(f'C2:C{rows + 1}')
    ws.add_data_validation(dv)
    dv = DataValidation(type='list', formula1=f'"{",".join(HAND_TYPE)}"', allow_blank=True)
    dv.add(f'D2:D{rows + 1}')
    ws.add_data_validation(dv)
    dv = DataValidation(type='list', formula1=f'"{",".join(LENGTH)}"', allow_blank=True)
    dv.add(f'E2:E{rows + 1}')
    ws.add_data_validation(dv)

    helper_col_idx = 6 + max_styles + 1 + 3
    helper_letter = get_column_letter(helper_col_idx)
    for i, v in enumerate(STYLE_OPTIONS, 1):
        ws.cell(row=i, column=helper_col_idx, value=v)
    ws.column_dimensions[helper_letter].hidden = True
    rng = f"'{sheet_name}'!${helper_letter}$1:${helper_letter}${len(STYLE_OPTIONS)}"
    for si in range(max_styles):
        col_letter = get_column_letter(6 + si)
        dv = DataValidation(type='list', formula1=f'={rng}', allow_blank=True)
        dv.add(f'{col_letter}2:{col_letter}{rows + 1}')
        ws.add_data_validation(dv)
    ws.freeze_panes = 'C2'


def build_hand_sheet(sheet_name, folder, labels):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    headers = ['序号', '图片', '手型']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center
        c.border = border
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.row_dimensions[1].height = 28
    for n, hand in enumerate(labels, 1):
        r = n + 1
        ws.row_dimensions[r].height = 95
        ws.cell(row=r, column=1, value=f'{n:02d}').alignment = align_center
        ws.cell(row=r, column=1).border = border
        add_image(ws, os.path.join(folder, f'{n:02d}.png'), r)
        ws.cell(row=r, column=2).border = border
        c = ws.cell(row=r, column=3, value=hand)
        c.alignment = align_center
        c.border = border
    rows = len(labels)
    dv = DataValidation(type='list', formula1=f'"{",".join(HAND_TYPE)}"', allow_blank=True)
    dv.add(f'C2:C{rows + 1}')
    ws.add_data_validation(dv)
    ws.freeze_panes = 'C2'


build_style_sheet('增强后款式图', '增强后款式图URL', labels)
build_style_sheet('款式图', '款式图URL', labels)
build_hand_sheet('手图', '手图URL', hand_labels)

wb.save(DST)
print(f'saved {DST}, size={os.path.getsize(DST)/1024/1024:.2f}MB')

wb2 = load_workbook(DST)
for n in ['增强后款式图', '款式图', '手图']:
    ws = wb2[n]
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1) if ws.cell(1, i).value]
    print(f'{n}: 图片={len(ws._images)}, 表头={headers}')
