import argparse
import csv
import re
import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl


SHEET_DICT = "\u7f8e\u7532\u9009\u62e9\u5206\u7c7b"
SHEET_STYLE = "\u6b3e\u5f0f\u56fe"
SHEET_ENHANCED = "\u589e\u5f3a\u540e\u6b3e\u5f0f\u56fe"
SHEET_HAND = "\u624b\u56fe"

DIR_ORIGINAL = "\u539f\u59cb\u6b3e\u5f0f\u56feURL"
DIR_ENHANCED = "\u589e\u5f3a\u540e\u6b3e\u5f0f\u56feURL"
DIR_STYLE_URL = "\u6b3e\u5f0f\u56feURL"
DIR_HAND = "\u624b\u56feURL"

TAG_NAIL_SHAPE = "\u7532\u578b"
TAG_HAND_TYPE = "\u624b\u578b"
TAG_LENGTH = "\u957f\u5ea6"
TAG_STYLE = "\u6b3e\u5f0f"
TAG_COLOR = "\u989c\u8272"
COL_RECOMMENDED_HAND = "\u63a8\u8350\u624b\u578b"

IMAGE_SUFFIXES = [".png", ".jpg", ".jpeg", ".webp", ".bmp"]


def clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def split_multi_value(text):
    text = clean_text(text)
    if not text:
        return []
    parts = re.split(r"[+\uFF0B\u3001,\uFF0C/\uFF0F;\uFF1B\s]+", text)
    return [part.strip() for part in parts if part.strip()]


def normalize_serial(value):
    text = clean_text(value)
    if not text:
        return ""
    if text.isdigit():
        return text.zfill(2)
    match = re.search(r"\d+", text)
    return match.group(0).zfill(2) if match else text


def find_image(folder, serial):
    folder = Path(folder)
    for suffix in IMAGE_SUFFIXES:
        candidate = folder / f"{serial}{suffix}"
        if candidate.exists():
            return candidate
    return None


def parse_style_sheet(workbook, sheet_name):
    ws = workbook[sheet_name]
    styles = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        serial = normalize_serial(row[0] if len(row) > 0 else "")
        if not serial or not serial.isdigit():
            continue
        style_tags = [clean_text(row[i]) for i in (5, 6, 7) if len(row) > i and clean_text(row[i])]
        styles[serial] = {
            "style_id": f"style_{serial}",
            "serial_no": serial,
            "nail_shape": clean_text(row[2] if len(row) > 2 else ""),
            "recommended_hand_type": clean_text(row[3] if len(row) > 3 else ""),
            "nail_length": clean_text(row[4] if len(row) > 4 else ""),
            "style_tags": style_tags,
            "color_text": clean_text(row[8] if len(row) > 8 else ""),
        }
    return styles


def parse_tag_dictionary(workbook):
    ws = workbook[SHEET_DICT]
    records = []
    current_category = ""
    current_secondary = ""
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        category = clean_text(row[0] if len(row) > 0 else "")
        secondary = clean_text(row[1] if len(row) > 1 else "")
        tag_name = clean_text(row[2] if len(row) > 2 else "")
        suitable = clean_text(row[3] if len(row) > 3 else "")
        unsuitable = clean_text(row[4] if len(row) > 4 else "")
        definition = clean_text(row[5] if len(row) > 5 else "")

        if category:
            current_category = category
            current_secondary = ""
        if secondary:
            current_secondary = secondary
        if not any([category, secondary, tag_name, suitable, unsuitable, definition]):
            continue

        records.append(
            {
                "source_row": idx,
                "category": current_category,
                "secondary_category": current_secondary,
                "tag_name": tag_name or secondary or category,
                "suitable_hand_types": suitable,
                "unsuitable_hand_types": unsuitable,
                "definition": definition,
            }
        )
    return records


def build_rows(excel_path, image_root):
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    original_styles = parse_style_sheet(workbook, SHEET_STYLE)
    enhanced_styles = parse_style_sheet(workbook, SHEET_ENHANCED)

    styles = []
    style_tags = []
    style_images = []
    issues = []

    all_serials = sorted(set(original_styles) | set(enhanced_styles))
    for serial in all_serials:
        original = original_styles.get(serial, {})
        enhanced = enhanced_styles.get(serial, {})
        source = original or enhanced
        style_id = f"style_{serial}"

        for key in ["nail_shape", "recommended_hand_type", "nail_length", "style_tags", "color_text"]:
            if original and enhanced and original.get(key) != enhanced.get(key):
                issues.append(f"style {serial}: {key} mismatch between style sheets")

        style_tag_names = source.get("style_tags", [])
        styles.append(
            {
                "style_id": style_id,
                "serial_no": serial,
                "nail_shape": source.get("nail_shape", ""),
                "recommended_hand_type": source.get("recommended_hand_type", ""),
                "nail_length": source.get("nail_length", ""),
                "primary_style": style_tag_names[0] if style_tag_names else "",
                "color_text": source.get("color_text", ""),
            }
        )

        fixed_tags = [
            (TAG_NAIL_SHAPE, source.get("nail_shape", ""), TAG_NAIL_SHAPE, 1.0),
            (TAG_HAND_TYPE, source.get("recommended_hand_type", ""), COL_RECOMMENDED_HAND, 1.0),
            (TAG_LENGTH, source.get("nail_length", ""), TAG_LENGTH, 1.0),
        ]
        for tag_type, tag_name, source_column, weight in fixed_tags:
            if tag_name:
                style_tags.append(
                    {
                        "style_id": style_id,
                        "tag_type": tag_type,
                        "tag_name": tag_name,
                        "source_column": source_column,
                        "weight": weight,
                    }
                )

        for order, tag_name in enumerate(style_tag_names, start=1):
            style_tags.append(
                {
                    "style_id": style_id,
                    "tag_type": TAG_STYLE,
                    "tag_name": tag_name,
                    "source_column": f"{TAG_STYLE}{order}",
                    "weight": max(0.6, 1.1 - order * 0.1),
                }
            )

        for color in split_multi_value(source.get("color_text", "")):
            style_tags.append(
                {
                    "style_id": style_id,
                    "tag_type": TAG_COLOR,
                    "tag_name": color,
                    "source_column": TAG_COLOR,
                    "weight": 0.8,
                }
            )

        for image_type, folder_name in [
            ("original", DIR_ORIGINAL),
            ("enhanced", DIR_ENHANCED),
            ("style_url", DIR_STYLE_URL),
        ]:
            folder = image_root / folder_name
            path = find_image(folder, serial)
            file_exists = path is not None
            style_images.append(
                {
                    "image_id": f"{style_id}_{image_type}",
                    "style_id": style_id,
                    "image_type": image_type,
                    "image_path": str(path) if file_exists else str(folder / f"{serial}.png"),
                    "file_exists": 1 if file_exists else 0,
                    "file_ext": path.suffix.lower() if file_exists else "",
                    "file_size": path.stat().st_size if file_exists else 0,
                }
            )
            if not file_exists and image_type in ("original", "enhanced"):
                issues.append(f"style {serial}: missing {image_type} image in {folder}")

    hand_images = []
    ws = workbook[SHEET_HAND]
    for row in ws.iter_rows(min_row=2, values_only=True):
        serial = normalize_serial(row[0] if len(row) > 0 else "")
        if not serial or not serial.isdigit():
            continue
        folder = image_root / DIR_HAND
        path = find_image(folder, serial)
        file_exists = path is not None
        hand_images.append(
            {
                "hand_id": f"hand_{serial}",
                "serial_no": serial,
                "hand_type": clean_text(row[2] if len(row) > 2 else ""),
                "image_path": str(path) if file_exists else str(folder / f"{serial}.png"),
                "file_exists": 1 if file_exists else 0,
                "file_ext": path.suffix.lower() if file_exists else "",
                "file_size": path.stat().st_size if file_exists else 0,
            }
        )
        if not file_exists:
            issues.append(f"hand {serial}: missing image in {folder}")

    return styles, style_tags, style_images, hand_images, parse_tag_dictionary(workbook), issues


def write_csv(path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def create_database(db_path, styles, style_tags, style_images, hand_images, tag_dictionary):
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        db_path.unlink()
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.executescript(
        """
        CREATE TABLE styles (
            style_id TEXT PRIMARY KEY,
            serial_no TEXT NOT NULL UNIQUE,
            nail_shape TEXT,
            recommended_hand_type TEXT,
            nail_length TEXT,
            primary_style TEXT,
            color_text TEXT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE style_images (
            image_id TEXT PRIMARY KEY,
            style_id TEXT NOT NULL,
            image_type TEXT NOT NULL,
            image_path TEXT NOT NULL,
            file_exists INTEGER NOT NULL DEFAULT 0,
            file_ext TEXT,
            file_size INTEGER,
            FOREIGN KEY(style_id) REFERENCES styles(style_id)
        );

        CREATE TABLE style_tags (
            tag_id INTEGER PRIMARY KEY AUTOINCREMENT,
            style_id TEXT NOT NULL,
            tag_type TEXT NOT NULL,
            tag_name TEXT NOT NULL,
            source_column TEXT,
            weight REAL NOT NULL DEFAULT 1.0,
            FOREIGN KEY(style_id) REFERENCES styles(style_id)
        );

        CREATE TABLE hand_images (
            hand_id TEXT PRIMARY KEY,
            serial_no TEXT NOT NULL UNIQUE,
            hand_type TEXT,
            image_path TEXT NOT NULL,
            file_exists INTEGER NOT NULL DEFAULT 0,
            file_ext TEXT,
            file_size INTEGER
        );

        CREATE TABLE tag_dictionary (
            tag_id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_row INTEGER,
            category TEXT,
            secondary_category TEXT,
            tag_name TEXT NOT NULL,
            suitable_hand_types TEXT,
            unsuitable_hand_types TEXT,
            definition TEXT
        );

        CREATE TABLE tryon_events (
            event_id TEXT PRIMARY KEY,
            user_id TEXT,
            hand_id TEXT,
            style_id TEXT,
            event_type TEXT NOT NULL,
            event_time TEXT NOT NULL,
            result_path TEXT,
            quality_score REAL,
            FOREIGN KEY(hand_id) REFERENCES hand_images(hand_id),
            FOREIGN KEY(style_id) REFERENCES styles(style_id)
        );

        CREATE VIEW style_tag_summary AS
        SELECT
            s.style_id,
            s.serial_no,
            s.nail_shape,
            s.recommended_hand_type,
            s.nail_length,
            s.primary_style,
            s.color_text,
            GROUP_CONCAT(t.tag_type || ':' || t.tag_name, '|') AS tags
        FROM styles s
        LEFT JOIN style_tags t ON s.style_id = t.style_id
        GROUP BY s.style_id;

        CREATE VIEW recommendation_base AS
        SELECT
            s.style_id,
            s.serial_no,
            s.recommended_hand_type,
            s.nail_shape,
            s.nail_length,
            s.primary_style,
            s.color_text,
            MAX(CASE WHEN i.image_type = 'enhanced' THEN i.image_path END) AS enhanced_image_path,
            MAX(CASE WHEN i.image_type = 'original' THEN i.image_path END) AS original_image_path
        FROM styles s
        LEFT JOIN style_images i ON s.style_id = i.style_id
        GROUP BY s.style_id;

        CREATE INDEX idx_style_tags_name ON style_tags(tag_type, tag_name);
        CREATE INDEX idx_styles_hand ON styles(recommended_hand_type);
        CREATE INDEX idx_tryon_events_style_time ON tryon_events(style_id, event_time);
        """
    )

    now = datetime.now().isoformat(timespec="seconds")
    conn.executemany(
        """
        INSERT INTO styles
        (style_id, serial_no, nail_shape, recommended_hand_type, nail_length, primary_style, color_text, created_at)
        VALUES (:style_id, :serial_no, :nail_shape, :recommended_hand_type, :nail_length, :primary_style, :color_text, :created_at)
        """,
        [{**row, "created_at": now} for row in styles],
    )
    conn.executemany(
        """
        INSERT INTO style_images
        (image_id, style_id, image_type, image_path, file_exists, file_ext, file_size)
        VALUES (:image_id, :style_id, :image_type, :image_path, :file_exists, :file_ext, :file_size)
        """,
        style_images,
    )
    conn.executemany(
        """
        INSERT INTO style_tags
        (style_id, tag_type, tag_name, source_column, weight)
        VALUES (:style_id, :tag_type, :tag_name, :source_column, :weight)
        """,
        style_tags,
    )
    conn.executemany(
        """
        INSERT INTO hand_images
        (hand_id, serial_no, hand_type, image_path, file_exists, file_ext, file_size)
        VALUES (:hand_id, :serial_no, :hand_type, :image_path, :file_exists, :file_ext, :file_size)
        """,
        hand_images,
    )
    conn.executemany(
        """
        INSERT INTO tag_dictionary
        (source_row, category, secondary_category, tag_name, suitable_hand_types, unsuitable_hand_types, definition)
        VALUES (:source_row, :category, :secondary_category, :tag_name, :suitable_hand_types, :unsuitable_hand_types, :definition)
        """,
        tag_dictionary,
    )
    conn.commit()
    return conn


def main():
    parser = argparse.ArgumentParser(description="Build normalized nail style database from Excel labels.")
    parser.add_argument("--excel", default=r"C:\Users\yzh\Desktop\美甲分类_打标.xlsx")
    parser.add_argument("--image-root", default=r"D:\AI_Project\nail\美甲图")
    parser.add_argument("--out-dir", default=r"D:\AI_Project\virtual-nail\data\style_database")
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    rows = build_rows(Path(args.excel), Path(args.image_root))
    styles, style_tags, style_images, hand_images, tag_dictionary, issues = rows
    conn = create_database(out_dir / "nail_style.db", styles, style_tags, style_images, hand_images, tag_dictionary)

    write_csv(out_dir / "styles.csv", styles, list(styles[0].keys()))
    write_csv(out_dir / "style_tags.csv", style_tags, list(style_tags[0].keys()))
    write_csv(out_dir / "style_images.csv", style_images, list(style_images[0].keys()))
    write_csv(out_dir / "hand_images.csv", hand_images, list(hand_images[0].keys()))
    write_csv(out_dir / "tag_dictionary.csv", tag_dictionary, list(tag_dictionary[0].keys()))
    (out_dir / "data_quality_report.txt").write_text(
        "\n".join(issues) if issues else "No blocking data quality issues found.",
        encoding="utf-8",
    )

    counts = {
        "styles": conn.execute("SELECT COUNT(*) FROM styles").fetchone()[0],
        "style_tags": conn.execute("SELECT COUNT(*) FROM style_tags").fetchone()[0],
        "style_images": conn.execute("SELECT COUNT(*) FROM style_images").fetchone()[0],
        "hand_images": conn.execute("SELECT COUNT(*) FROM hand_images").fetchone()[0],
        "tag_dictionary": conn.execute("SELECT COUNT(*) FROM tag_dictionary").fetchone()[0],
        "missing_images": conn.execute("SELECT COUNT(*) FROM style_images WHERE file_exists = 0").fetchone()[0],
    }
    conn.close()

    print(f"database: {out_dir / 'nail_style.db'}")
    for key, value in counts.items():
        print(f"{key}: {value}")
    print(f"csv_dir: {out_dir}")
    print(f"quality_report: {out_dir / 'data_quality_report.txt'}")


if __name__ == "__main__":
    main()
